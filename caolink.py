# Import libraries and packages for the project 

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from bs4 import BeautifulSoup
from time import sleep
from openpyxl import load_workbook
import openpyxl
import time
import random
print('- Finish importing packages')
#Gọi đăng nhập
def loginbds():
    # Task 1.1: Open Chrome and Access Linkedin login site
    sleep(1)
    urldn = 'https://batdongsan.com.vn/'
    driver.get(urldn)
    print('- hàm gọi dang nhap')
    sleep(1)
    # Task 1.2: Import username and password
    credential = open('credentials.txt')
    line = credential.readlines()
    username = line[0]
    password = line[1]
    sleep(1)

    # Task 1.2: Key in login credentials
    dangnhap = driver.find_element_by_xpath("/html/body/div[1]/header/div[2]/div[1]/div[2]/a[1]")
    dangnhap.click()
    sleep(2)
    email_field = driver.find_element_by_id('UserName')
    email_field.send_keys(username)
    print('- Finish keying in email')
    sleep(1)

    password_field = driver.find_element_by_id('Password')
    password_field.send_keys(password)
    print('- Finish keying in password')
    sleep(2)

    # Task 1.2: Click the Login button
    signin_field = driver.find_element_by_id('btnLogin')
    signin_field.click()
    sleep(3)
    print('- Finish Task 1: Login to bds')
# setup selenium webdriver
opt = webdriver.ChromeOptions()
prefs = {"profile.managed_default_content_settings.images": 2}
opt.add_experimental_option("prefs", prefs)
opt.add_argument('--disable-gpu')
opt.add_argument("--window-size=1920,1080")
opt.add_argument("--start-maximized")
opt.add_argument('--disable-dev-shm-usage')
opt.add_argument('--no-sandbox')
opt.add_argument('--disable-notifications')
opt.add_argument('--ignore-certificate-errors')
opt.add_argument("--disable-browser-side-navigation")
# opt.add_argument('--headless')
opt.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36")
driver = webdriver.Chrome(options=opt)
#login bds
loginbds()
sleep(2)
url = 'https://batdongsan.com.vn/cho-thue-can-ho-chung-cu-tp-hcm'
driver.get(url)
sleep(2)

print('vào website')

def geturl():
    page_source = BeautifulSoup(driver.page_source,"html.parser")
    # print(page_source)
    urls_bds = page_source.find_all("a",class_="wrap-plink")
    all_URL = []
    for url_bds in urls_bds:
        url_ID = url_bds.get('href')
        url_URL = "https://batdongsan.com.vn" + url_ID
        if url_URL not in all_URL:
            all_URL.append(url_URL)
    return(all_URL)

input_page = int(input('How many pages you want to scrape: '))
URLs_all_page = []

for page in range(input_page):
    URLs_one_page = geturl()
    print(f"đã cào trang thứ {page+1}")
    sleep(2)
    a = page+2
    nexturl = url+"/p" + str(a)
    # print(nexturl)
    driver.get(nexturl)
    URLs_all_page = URLs_all_page + URLs_one_page
    sleep(2)
    # if a < input_page:
    #     nexturl = url+"/p" + str(a)
    # # print(nexturl)
    #     driver.get(nexturl)
    #     URLs_all_page = URLs_all_page + URLs_one_page
    #     sleep(2)
print('- Finish Task 3: Scrape the URLs')
#GHi số vào file
b =2
for linkbds in URLs_all_page:
    wa = load_workbook('ketquabds.xlsx')
    was = wa.worksheets[0]
    # print(linkbds)
    tdlink1 = "d" + str(b)
    # print(tdlink1)
    was[tdlink1] = linkbds
    b = b + 1
    wa.save('ketquabds.xlsx')
    wa.close()
    # print("có vào lưu file")

print("lưu link rồi")

i= 2
for linkurl in URLs_all_page:
    #quanr li proxy
    driver.get(linkurl)
    # print('- Accessing profile: ', linkctcao)
    page_source = BeautifulSoup(driver.page_source, "html.parser")
    info_div = page_source.find('div',{'class':'user'})
    # time.sleep(random.randint(2, 3))
    try:
        info_loc = info_div.find_all('div')
        # print(info_loc)
        try:
            name = info_loc[1].get("title")
        except:
            name ="không có"
        #HÀM THAY ĐỔI    
        # print(info_loc[2])
        ploai = info_loc[2].get("class")
        # print(type(ploai))
        if ploai == ["info"]:
            sdtraw = info_loc[3].find("span")
            try:
                sdt= sdtraw.get("raw")
            except:
                sdt="không có"
            emailraw = info_loc[4].find("a")
            try:
                email =emailraw.get("data-email")
            except:
                email="không có"
        else:
            sdtraw = info_loc[2].find("span")
            try:
                sdt= sdtraw.get("raw")
            except:
                sdt="không có"
            emailraw = info_loc[3].find("a")
            try:
                email =emailraw.get("data-email")
            except:
                email="không có"
        print(f"thông tin khách hàng tên {name} sdt là {sdt} có email là {email}")
        tdten = "a"+ str(i)
        tdsdt = "b"+ str(i)
        tdemail = "c" + str(i)
        # tdlink = "d" + str(i)
        wa = load_workbook('ketquabds.xlsx')
        activesheet = wa.worksheets[0]
        activesheet[tdten] = name
        activesheet[tdsdt] = sdt
        activesheet[tdemail] = email
        # activesheet[tdlink] = linkurl
        i = i +1
        wa.save('ketquabds.xlsx')
        wa.close()
        # time.sleep(random.randint(2, 3))
    except:
        pass
        i = i +1
    print(f"Đã xử lí xong {i-1} khách hàng")
driver.quit()
print('Mission Completed!')

