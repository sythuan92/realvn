from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from bs4 import BeautifulSoup
from time import sleep
from openpyxl import load_workbook
import openpyxl
import time
import random


wb = load_workbook('ketquabds.xlsx')
activesheet = wb.worksheets[0]
row_count = activesheet.max_row
print(row_count)


# proxy = "117.0.161.170:14514"
# setup selenium webdriver
opt = webdriver.ChromeOptions()
# opt.add_extension("Block-image_v1.1.crx")
prefs = {"profile.managed_default_content_settings.images": 2}
opt.add_experimental_option("prefs", prefs)
opt.add_argument('--disable-gpu')
opt.add_argument("--window-size=1920,1080")
opt.add_argument("--start-maximized")
opt.add_argument('--disable-dev-shm-usage')
opt.add_argument('--no-sandbox')
opt.add_argument('--ignore-certificate-errors')
opt.add_argument("--disable-browser-side-navigation")
# opt.add_argument('--headless')
# opt.add_argument('--proxy-server={}'.format(proxy))
opt.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36")
# driver = webdriver.Chrome(options=opt)
# driver = webdriver.Chrome(options=opt)


i = 70 # phai bat dau tu 2 vi 1 la tieu de
for linkurl in range(row_count):
    #quanr li proxy
    if i % 40 == 0:
        proxy = input('thay proxy mới đi ông:')
    else:
        pass
    # opt.add_argument('--proxy-server={}'.format(proxy))
    driver = webdriver.Chrome(options=opt)
    #thuc hien cao link
    wb = load_workbook('ketquabds.xlsx')
    activesheet = wb.worksheets[0]
    linkvitri = "d" + str(i)
    linkctcao = activesheet[linkvitri].value
    # print(f"in linkctcao {linkctcao}")
    driver.get(linkctcao)
    # print('- Accessing profile: ', linkctcao)
    page_source = BeautifulSoup(driver.page_source, "html.parser")
    info_div = page_source.find('div',{'class':'user'})
    # time.sleep(random.randint(2, 3))
    try:
        info_loc = info_div.find_all('div')
        # print(info_loc)
        try:
            name = info_loc[1].get("title")
        except:
            name ="không có"
        #HÀM THAY ĐỔI    
        # print(info_loc[2])
        ploai = info_loc[2].get("class")
        # print(type(ploai))
        if ploai == ["info"]:
            sdtraw = info_loc[3].find("span")
            try:
                sdt= sdtraw.get("raw")
            except:
                sdt="không có"
            emailraw = info_loc[4].find("a")
            try:
                email =emailraw.get("data-email")
            except:
                email="không có"
        else:
            sdtraw = info_loc[2].find("span")
            try:
                sdt= sdtraw.get("raw")
            except:
                sdt="không có"
            emailraw = info_loc[3].find("a")
            try:
                email =emailraw.get("data-email")
            except:
                email="không có"
        # print(f"thông tin khách hàng tên {name} sdt là {sdt} có email là {email}")
        tdten = "a"+ str(i)
        tdsdt = "b"+ str(i)
        tdemail = "c" + str(i)
        # tdlink = "d" + str(i)
        activesheet = wb.worksheets[0]
        activesheet[tdten] = name
        activesheet[tdsdt] = sdt
        activesheet[tdemail] = email
        # activesheet[tdlink] = linkedin_URL
        i = i +1
        wb.save('ketquabds.xlsx')
        # time.sleep(random.randint(2, 3))
    except:
        pass
    i = i +1
    driver.quit()
    print(f"Đã xử lí xong {i-1} khách hàng")

print('Mission Completed!')